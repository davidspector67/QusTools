from UtcTool2d.exportData_ui import *
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import re
from PyQt5.QtWidgets import QWidget, QApplication, QFileDialog

import platform
system = platform.system()

class ExportDataGUI(Ui_exportData, QWidget):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        if system == 'Windows':
            self.roiSidebarLabel.setStyleSheet("""QLabel { 
                font-size: 18px; 
                color: rgb(255, 255, 255); 
                background-color: rgba(255, 255, 255, 0); 
                border: 0px; 
                font-weight: bold; 
            }""")
            self.imageSelectionLabelSidebar.setStyleSheet("""QLabel {
                font-size: 18px;
                color: rgb(255, 255, 255);
                background-color: rgba(255, 255, 255, 0);
                border: 0px;
                font-weight: bold;
            }""")
            self.imageLabel.setStyleSheet("""QLabel {
                font-size: 13px;
                color: rgb(255, 255, 255);
                background-color: rgba(255, 255, 255, 0);
                border: 0px;
                font-weight: bold;
            }""")
            self.phantomLabel.setStyleSheet("""QLabel {
                font-size: 13px;
                color: rgb(255, 255, 255);
                background-color: rgba(255, 255, 255, 0);
                border: 0px;
                font-weight: bold;
            }""")
            self.imagePathInput.setStyleSheet("""QLabel {
                font-size: 11px;
                color: rgb(255, 255, 255);
                background-color: rgba(255, 255, 255, 0);
                border: 0px;
            }""")
            self.phantomPathInput.setStyleSheet("""QLabel {
                font-size: 11px;
                color: rgb(255, 255, 255);
                background-color: rgba(255, 255, 255, 0);
                border: 0px;
            }""")
            self.analysisParamsLabel.setStyleSheet("""QLabel {
                font-size: 18px;
                color: rgb(255, 255, 255);
                background-color: rgba(255, 255, 255, 0);
                border: 0px;
                font-weight:bold;
            }""")
            self.rfAnalysisLabel.setStyleSheet("""QLabel {
                font-size: 18px;
                color: rgb(255, 255, 255);
                background-color: rgba(255, 255, 255, 0);
                border: 0px;
                font-weight:bold;
            }""")
            self.exportResultsLabel.setStyleSheet("""QLabel {
                font-size: 18px;
                color: rgb(255, 255, 255);
                background-color: rgba(255, 255, 255, 0);
                border: 0px;
                font-weight:bold;
            }""")

        self.newFolderPathInput.setHidden(True)
        self.newFileNameInput.setHidden(True)
        self.newFolderPathLabel.setHidden(True)
        self.newFileNameLabel.setHidden(True)
        self.chooseAppendFileButton.setHidden(True)
        self.chooseNewFolderButton.setHidden(True)
        self.appendFileButton.setHidden(True)
        self.createNewFileButton.setHidden(True)
        self.appendFileLabel.setHidden(True)
        self.appendFilePath.setHidden(True)
        self.clearAppendFileButton.setHidden(True)
        self.clearNewFolderButton.setHidden(True)
        self.appendFileBackButton.setHidden(True)
        self.newFileBackButton.setHidden(True)
        self.fileNameErrorLabel.setHidden(True)
        self.fileNameWarningLabel.setHidden(True)
        self.dataSavedSuccessfullyLabel.setHidden(True)

        self.dataFrame = None
        self.lastGui = None

        self.newFileOptionButton.clicked.connect(self.newFileOptionSelected)
        self.newFileBackButton.clicked.connect(self.backFromNewFileOption)
        self.appendFileOptionButton.clicked.connect(self.appendOptionSelected)
        self.appendFileBackButton.clicked.connect(self.backFromAppendOption)
        self.backButton.clicked.connect(self.backToAnalysis)
        self.appendFileButton.clicked.connect(self.appendToFile)
        self.createNewFileButton.clicked.connect(self.createNewFile)
        self.chooseNewFolderButton.clicked.connect(self.selectNewFolder)
        self.chooseAppendFileButton.clicked.connect(self.selectExistingFile)
        self.clearNewFolderButton.clicked.connect(self.clearNewFolder)
        self.clearAppendFileButton.clicked.connect(self.clearNewFile)
    
    def clearNewFolder(self):
        self.newFolderPathInput.clear()

    def setFilenameDisplays(self, imageName, phantomName):
        self.imagePathInput.setHidden(False)
        self.phantomPathInput.setHidden(False)
        self.imagePathInput.setText(imageName)
        self.phantomPathInput.setText(phantomName)
    
    def clearNewFile(self):
        self.appendFilePath.clear()

    def selectNewFolder(self):
        folderName = QFileDialog.getExistingDirectory(None, 'Select Directory')
        if folderName != '':
            self.newFolderPathInput.setText(folderName)

    def selectExistingFile(self):
        fileName, _ = QFileDialog.getOpenFileName(None, 'Open file', filter = '*.xlsx')
        if fileName != '':
            self.appendFilePath.setText(fileName)

    def dataSavedSuccessfully(self):
        self.dataSavedSuccessfullyLabel.setHidden(False)
        self.newFolderPathInput.setHidden(True)
        self.newFileNameInput.setHidden(True)
        self.newFolderPathLabel.setHidden(True)
        self.newFileNameLabel.setHidden(True)
        self.chooseAppendFileButton.setHidden(True)
        self.chooseNewFolderButton.setHidden(True)
        self.appendFileButton.setHidden(True)
        self.createNewFileButton.setHidden(True)
        self.appendFileLabel.setHidden(True)
        self.appendFilePath.setHidden(True)
        self.clearAppendFileButton.setHidden(True)
        self.clearNewFolderButton.setHidden(True)
        self.appendFileBackButton.setHidden(True)
        self.newFileBackButton.setHidden(True)
        self.fileNameErrorLabel.setHidden(True)
        self.fileNameWarningLabel.setHidden(True)
        self.appendFileOptionButton.setHidden(True)
        self.newFileOptionButton.setHidden(True)
        self.selectDataLabel.setHidden(True)

    def createNewFile(self):
        if os.path.exists(self.newFolderPathInput.text()):
            if not (self.newFileNameInput.text().endswith(".xlsx") and (not bool(re.search(r"\s", self.newFileNameInput.text())))):
                self.fileNameWarningLabel.setHidden(True)
                self.fileNameErrorLabel.setHidden(False)
                return
            try:
                wb = Workbook()
                ws = wb.active
                for r in dataframe_to_rows(self.dataFrame, index=False, header=True):
                    ws.append(r)
                wb.save(os.path.join(self.newFolderPathInput.text(), self.newFileNameInput.text()))
                wb.close()
                
                self.dataSavedSuccessfully()
            except Exception as e:
                print(str(e))


    def appendToFile(self):
        if os.path.exists(self.appendFilePath.text()) and self.appendFilePath.text().endswith(".xlsx"):
            try:
                # Since writes to 'Sheet1', make sure not to change sheet names
                wb = load_workbook(self.appendFilePath.text())
                ws = wb.active
                for r in dataframe_to_rows(self.dataFrame, index=False, header=False):
                    ws.append(r)
                wb.save(self.appendFilePath.text())
                wb.close()

                self.dataSavedSuccessfully()
            except Exception as e:
                print(str(e))

    def backToAnalysis(self):
        self.lastGui.show()
        self.hide()

    def newFileOptionSelected(self):
        self.chooseNewFolderButton.setHidden(False)
        self.newFolderPathInput.setHidden(False)
        self.newFolderPathLabel.setHidden(False)
        self.clearNewFolderButton.setHidden(False)
        self.chooseNewFolderButton.setHidden(False)
        self.newFileNameInput.setHidden(False)
        self.newFileNameLabel.setHidden(False)
        self.newFileBackButton.setHidden(False)
        self.createNewFileButton.setHidden(False)
        self.fileNameWarningLabel.setHidden(False)
        self.appendFileOptionButton.setHidden(True)
        self.newFileOptionButton.setHidden(True)

    def backFromNewFileOption(self):
        self.chooseNewFolderButton.setHidden(True)
        self.newFolderPathInput.setHidden(True)
        self.newFolderPathLabel.setHidden(True)
        self.clearNewFolderButton.setHidden(True)
        self.chooseNewFolderButton.setHidden(True)
        self.newFileNameInput.setHidden(True)
        self.newFileNameLabel.setHidden(True)
        self.createNewFileButton.setHidden(True)
        self.newFileBackButton.setHidden(True)
        self.fileNameWarningLabel.setHidden(True)
        self.fileNameErrorLabel.setHidden(True)
        self.appendFileOptionButton.setHidden(False)
        self.newFileOptionButton.setHidden(False)

    def appendOptionSelected(self):
        self.chooseAppendFileButton.setHidden(False)
        self.appendFileButton.setHidden(False)
        self.appendFileLabel.setHidden(False)
        self.appendFilePath.setHidden(False)
        self.clearAppendFileButton.setHidden(False)
        self.appendFileBackButton.setHidden(False)
        self.appendFileOptionButton.setHidden(True)
        self.newFileOptionButton.setHidden(True)

    def backFromAppendOption(self):
        self.chooseAppendFileButton.setHidden(True)
        self.appendFileButton.setHidden(True)
        self.appendFileLabel.setHidden(True)
        self.appendFilePath.setHidden(True)
        self.clearAppendFileButton.setHidden(True)
        self.appendFileBackButton.setHidden(True)
        self.appendFileOptionButton.setHidden(False)
        self.newFileOptionButton.setHidden(False)

if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    # selectWindow = QWidget()
    ui = ExportDataGUI()
    # ui.selectImage.show()
    ui.show()
    sys.exit(app.exec_())
    